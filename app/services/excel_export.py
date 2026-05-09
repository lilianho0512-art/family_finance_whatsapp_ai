from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from app.config import settings
from app.models import Family
from app.services import record_service, loan_service
from app.utils import fx
from app.utils.date_tools import month_range

HEADERS = [
    "ID", "Date", "Type", "Merchant", "Amount", "Currency", "Category",
    "Payment Method", "Source", "Note", "Status", "WhatsApp",
]


def _row(r):
    return [
        r.id,
        r.date.isoformat() if r.date else "",
        r.record_type,
        r.merchant or "",
        float(r.amount or 0),
        r.currency or "MYR",
        r.category or "",
        r.payment_method or "",
        r.source or "",
        (r.note or "")[:200],
        r.status or "",
        r.whatsapp_number or "",
    ]


def _bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)


def export_monthly(db: Session, family_id: Optional[int], ref: date = None) -> str:
    if ref is None:
        ref = date.today()
    s, e = month_range(ref)
    wb = Workbook()

    fam = db.query(Family).get(family_id) if family_id else None
    base_cur = (fam.default_currency if fam else "MYR") or "MYR"

    inc_g = record_service.month_total_grouped(db, family_id, "income", ref)
    exp_g = record_service.month_total_grouped(db, family_id, "expense", ref)
    sav_g = record_service.month_total_grouped(db, family_id, "savings", ref)
    inc = fx.convert_grouped(inc_g, base_cur)
    exp = fx.convert_grouped(exp_g, base_cur)
    sav = fx.convert_grouped(sav_g, base_cur)
    rate = round((sav / inc * 100), 2) if inc > 0 else 0.0
    has_fx = any(len(g) > 1 or (g and base_cur not in g) for g in (inc_g, exp_g, sav_g))

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Family Finance Monthly Report"])
    ws.append(["Period", f"{s.isoformat()} ~ {e.isoformat()}"])
    ws.append(["Family ID", family_id if family_id is not None else "(all)"])
    ws.append(["Reporting currency", base_cur])
    ws.append([])
    ws.append(["Income",  inc])
    ws.append(["Expense", exp])
    ws.append(["Savings", sav])
    ws.append(["Savings Rate %", rate])
    if has_fx:
        ws.append([])
        ws.append(["Native breakdown (pre-conversion)"])
        for label, g in (("Income", inc_g), ("Expense", exp_g), ("Savings", sav_g)):
            for cur, amt in (g or {}).items():
                ws.append([f"  {label} ({cur})", amt])
    ws["A1"].font = Font(bold=True, size=14)

    all_records = record_service.list_all(db, family_id)

    def add_sheet(name, rtype):
        ws = wb.create_sheet(name)
        ws.append(HEADERS)
        _bold_header(ws)
        for r in all_records:
            if r.record_type == rtype and r.date and s <= r.date <= e:
                ws.append(_row(r))

    add_sheet("Expenses", "expense")
    add_sheet("Savings", "savings")
    add_sheet("Income", "income")

    ws_cat = wb.create_sheet("Category Breakdown")
    ws_cat.append(["Category", "Total"])
    _bold_header(ws_cat)
    for cat, total in record_service.category_breakdown(db, family_id, ref):
        ws_cat.append([cat, total])

    ws_cf = wb.create_sheet("Cashflow")
    ws_cf.append(["Type", "Amount"])
    _bold_header(ws_cf)
    ws_cf.append(["Income", inc])
    ws_cf.append(["Expense", exp])
    ws_cf.append(["Savings", sav])
    ws_cf.append(["Net (Income - Expense - Savings)", inc - exp - sav])

    ws_nr = wb.create_sheet("Need Review")
    ws_nr.append(HEADERS + ["Missing Fields"])
    _bold_header(ws_nr)
    for r in all_records:
        if r.status in ("need_review", "need_question", "failed"):
            ws_nr.append(_row(r) + [r.missing_fields or ""])

    ws_l = wb.create_sheet("Loans")
    ws_l.append([
        "ID", "Kind", "Lender", "Currency", "Principal", "Current Balance", "Monthly Payment",
        "Interest Rate %", "Term (months)", "Start Date", "Due Day", "Status", "Notes",
    ])
    _bold_header(ws_l)
    if family_id is not None:
        loans = loan_service.list_loans(db, family_id)
        for l in loans:
            ws_l.append([
                l.id,
                l.kind,
                l.lender,
                l.currency or "MYR",
                float(l.principal or 0),
                float(l.current_balance or 0),
                float(l.monthly_payment or 0),
                l.interest_rate if l.interest_rate is not None else "",
                l.term_months or "",
                l.start_date.isoformat() if l.start_date else "",
                l.payment_due_day or "",
                l.status,
                (l.notes or "")[:200],
            ])
        ws_l.append([])
        ws_l.append(["Total monthly payment (active)", "", "", "", "", "", loan_service.total_monthly_payment(db, family_id)])
        ws_l.append(["Total outstanding (active)",     "", "", "", "", loan_service.total_outstanding(db, family_id)])

    settings.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fam_tag = f"_fam{family_id}" if family_id is not None else ""
    out_path = settings.OUTPUT_DIR / f"monthly_finance_report_{s.strftime('%Y%m')}{fam_tag}.xlsx"
    wb.save(out_path)
    return str(out_path)
