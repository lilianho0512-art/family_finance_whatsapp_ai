import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.services import rule_parser, menu_service, question_engine
from app.utils.json_tools import extract_json
from app.utils.money_tools import extract_amount


def test_greeting():
    assert menu_service.is_greeting("Hi")
    assert menu_service.is_greeting("hello")
    assert menu_service.is_greeting("Menu")
    assert not menu_service.is_greeting("Today Tesco RM88")


def test_rule_parse_expense():
    r = rule_parser.parse("Today Tesco RM88")
    assert r["amount"] == 88.0
    assert r["merchant"].lower() == "tesco"


def test_rule_parse_savings():
    r = rule_parser.parse("Today saved RM500")
    assert r["record_type"] == "savings"
    assert r["amount"] == 500.0


def test_rule_parse_income():
    r = rule_parser.parse("Salary RM3800")
    assert r["record_type"] == "income"
    assert r["amount"] == 3800.0
    assert r["source"] == "Salary"


def test_rule_parse_groceries():
    r = rule_parser.parse("Today groceries RM35")
    assert r["amount"] == 35.0
    assert r["category"] == "Groceries"


def test_extract_amount():
    assert extract_amount("RM 88.50") == 88.5
    assert extract_amount("MYR123") == 123.0
    assert extract_amount("Total 99") == 99.0


def test_extract_json_from_markdown():
    raw = '```json\n{"a": 1, "b": "x"}\n```'
    assert extract_json(raw) == {"a": 1, "b": "x"}


def test_extract_json_with_trailing_comma():
    raw = '{"a": 1, "b": 2,}'
    assert extract_json(raw) == {"a": 1, "b": 2}


def test_question_engine_letters():
    assert question_engine.resolve_answer("ask_record_type", "A") == "expense"
    assert question_engine.resolve_answer("ask_category", "A") == "Groceries"
    assert question_engine.resolve_answer("ask_payment_method", "D") == "Touch n Go"
    assert question_engine.resolve_answer("ask_record_type", "Z") is None


def test_explicit_type_prefixes():
    # Mixed case, since detect_record_type lowercases input
    assert rule_parser.detect_record_type("family expense RM50") == "expense"
    assert rule_parser.detect_record_type("Family Expense Tesco RM50") == "expense"
    assert rule_parser.detect_record_type("family savings RM500") == "savings"
    assert rule_parser.detect_record_type("family income RM3800") == "income"
    assert rule_parser.detect_record_type("family transfer RM200") == "transfer"


def test_transfer_standalone_vs_payment_method():
    # "transfer" alone => transfer record
    assert rule_parser.detect_record_type("transfer RM200 to mom") == "transfer"
    # "bank transfer" / "online transfer" are payment methods, not transfer records
    assert rule_parser.detect_record_type("Tesco RM50 bank transfer") != "transfer"
    assert rule_parser.detect_record_type("Tesco RM50 online transfer") != "transfer"
    # DuitNow keyword
    assert rule_parser.detect_record_type("duitnow RM150 to dad") == "transfer"


def test_query_patterns_match():
    from app.routers.whatsapp import _QUERY_PATTERNS

    def kind_for(text):
        for rx, kind in _QUERY_PATTERNS:
            if rx.search(text):
                return kind
        return None

    assert kind_for("how much did I spend this month") == "month_expense"
    assert kind_for("this month savings") == "month_savings"
    assert kind_for("this month income") == "month_income"
    assert kind_for("savings rate") == "savings_rate"
    assert kind_for("today expense") == "today_expense"
    assert kind_for("monthly summary") == "month_summary"
    assert kind_for("export") == "export"
    assert kind_for("hello there") is None


def test_undo_and_delete_regexes():
    from app.routers.whatsapp import _UNDO_RE, _DELETE_BY_ID_RE
    assert _UNDO_RE.match("undo")
    assert _UNDO_RE.match("/undo")
    assert _UNDO_RE.match("delete last")
    assert _UNDO_RE.match("delete recent")
    assert not _UNDO_RE.match("undo this expense")  # only bare command
    m = _DELETE_BY_ID_RE.match("delete #5")
    assert m and m.group(1) == "5"
    m = _DELETE_BY_ID_RE.match("/del 12")
    assert m and m.group(1) == "12"
    assert not _DELETE_BY_ID_RE.match("delete the cat")


def test_account_command_regexes():
    from app.routers.whatsapp import _BALANCE_LIST_RE, _BALANCE_SET_RE, _ADD_ACCOUNT_RE
    assert _BALANCE_LIST_RE.match("balance")
    assert _BALANCE_LIST_RE.match("Balances")
    assert not _BALANCE_LIST_RE.match("balance Maybank")
    m = _BALANCE_SET_RE.match("set Maybank 5000")
    assert m and m.group(1).strip() == "Maybank" and float(m.group(2)) == 5000.0
    m = _BALANCE_SET_RE.match("set Hong Leong RM 1234.50")
    assert m and m.group(1).strip() == "Hong Leong" and float(m.group(2)) == 1234.5
    m = _ADD_ACCOUNT_RE.match("add account OCBC")
    assert m and m.group(1).strip() == "OCBC"
    assert not _ADD_ACCOUNT_RE.match("add OCBC")


# ---------------------------------------------------------------------------
# Step-by-step flow (ask_amount step)
# ---------------------------------------------------------------------------


def test_resolve_amount_answer():
    assert question_engine.resolve_answer("ask_amount", "88.50") == 88.5
    assert question_engine.resolve_answer("ask_amount", "RM 1234.50") == 1234.5
    assert question_engine.resolve_answer("ask_amount", "MYR123") == 123.0
    assert question_engine.resolve_answer("ask_amount", "abc") is None
    assert question_engine.resolve_answer("ask_amount", "0") is None


class _StubRecord:
    def __init__(self, **kw):
        self.record_type = kw.get("record_type", "unknown")
        self.amount = kw.get("amount", 0)
        self.currency = kw.get("currency", "MYR")
        self.category = kw.get("category", "")
        self.payment_method = kw.get("payment_method", "")
        self.source = kw.get("source", "")
        self.account = kw.get("account", "")


def test_determine_next_asks_amount_after_record_type():
    # Record_type known but amount missing -> ask_amount comes next
    rec = _StubRecord(record_type="expense", amount=0)
    nxt = question_engine.determine_next_question(rec)
    assert nxt is not None
    step, _q, _opts = nxt
    assert step == "ask_amount"


def test_determine_next_skips_amount_when_present():
    rec = _StubRecord(record_type="expense", amount=50.0)
    step, _q, _opts = question_engine.determine_next_question(rec)
    assert step == "ask_category"  # already past ask_amount


def test_determine_next_unknown_type_first():
    rec = _StubRecord(record_type="unknown", amount=0)
    step, _q, _opts = question_engine.determine_next_question(rec)
    assert step == "ask_record_type"  # type still trumps amount


# ---------------------------------------------------------------------------
# Loan service CRUD + family scoping
# ---------------------------------------------------------------------------


def _make_db():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app.database import Base
    import app.models  # ensure models register on Base
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    return Session()


def _make_family(db, name="TestFam"):
    from app.models import Family
    f = Family(name=name)
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


def test_loan_create_and_list():
    from app.services import loan_service
    from datetime import date as _date
    db = _make_db()
    fam = _make_family(db)
    loan = loan_service.create_loan(
        db, fam.id,
        lender="Maybank Home Loan",
        principal=300000.0,
        monthly_payment=1500.0,
        interest_rate=4.25,
        term_months=360,
        start_date=_date(2026, 1, 1),
        payment_due_day=10,
        notes="30-yr mortgage",
    )
    assert loan.id is not None
    assert loan.kind == "loan"
    assert loan.status == "active"
    assert loan.current_balance == 300000.0  # defaults to principal
    listed = loan_service.list_loans(db, fam.id)
    assert len(listed) == 1 and listed[0].id == loan.id


def test_loan_kind_installment():
    from app.services import loan_service
    db = _make_db()
    fam = _make_family(db)
    plan = loan_service.create_loan(
        db, fam.id,
        lender="Shopee SPayLater",
        principal=1200.0,
        monthly_payment=400.0,
        kind="installment",
    )
    assert plan.kind == "installment"
    # Invalid kind falls back to "loan"
    bad = loan_service.create_loan(
        db, fam.id, lender="x", principal=10, monthly_payment=10, kind="garbage"
    )
    assert bad.kind == "loan"


def test_loan_family_scoping():
    from app.services import loan_service
    db = _make_db()
    fam_a = _make_family(db, "A")
    fam_b = _make_family(db, "B")
    a_loan = loan_service.create_loan(db, fam_a.id, lender="A Bank", principal=100, monthly_payment=10)
    loan_service.create_loan(db, fam_b.id, lender="B Bank", principal=200, monthly_payment=20)
    # Family A only sees their own
    a_loans = loan_service.list_loans(db, fam_a.id)
    assert len(a_loans) == 1 and a_loans[0].id == a_loan.id
    # Cross-family get returns None
    assert loan_service.get_loan(db, fam_b.id, a_loan.id) is None
    # Cross-family delete fails
    assert loan_service.delete_loan(db, fam_b.id, a_loan.id) is False
    # Loan still exists in A
    assert loan_service.get_loan(db, fam_a.id, a_loan.id) is not None


def test_loan_update_close_delete():
    from app.services import loan_service
    db = _make_db()
    fam = _make_family(db)
    loan = loan_service.create_loan(db, fam.id, lender="X", principal=1000, monthly_payment=100)
    # Update
    upd = loan_service.update_loan(db, fam.id, loan.id, monthly_payment=120, current_balance=900)
    assert upd.monthly_payment == 120 and upd.current_balance == 900
    # Close
    closed = loan_service.close_loan(db, fam.id, loan.id)
    assert closed.status == "closed" and closed.current_balance == 0.0
    # Delete
    assert loan_service.delete_loan(db, fam.id, loan.id) is True
    assert loan_service.get_loan(db, fam.id, loan.id) is None


def test_loan_totals():
    from app.services import loan_service
    db = _make_db()
    fam = _make_family(db)
    loan_service.create_loan(db, fam.id, lender="A", principal=500, monthly_payment=50)
    loan_service.create_loan(db, fam.id, lender="B", principal=1000, monthly_payment=100, current_balance=800)
    closed = loan_service.create_loan(db, fam.id, lender="C", principal=300, monthly_payment=30)
    loan_service.close_loan(db, fam.id, closed.id)
    # Totals only count active loans
    assert loan_service.total_monthly_payment(db, fam.id) == 150.0
    assert loan_service.total_outstanding(db, fam.id) == 1300.0  # 500 default + 800 set


def test_loan_due_day_validation():
    from app.services import loan_service
    db = _make_db()
    fam = _make_family(db)
    loan = loan_service.create_loan(
        db, fam.id, lender="X", principal=100, monthly_payment=10, payment_due_day=99
    )
    assert loan.payment_due_day is None  # invalid -> nulled
    upd = loan_service.update_loan(db, fam.id, loan.id, payment_due_day=15)
    assert upd.payment_due_day == 15


# ---------------------------------------------------------------------------
# Excel export Loans sheet
# ---------------------------------------------------------------------------


def test_excel_export_includes_loans_sheet(tmp_path, monkeypatch):
    from openpyxl import load_workbook
    from app.services import excel_export, loan_service
    from app import config as app_config
    db = _make_db()
    fam = _make_family(db)
    loan_service.create_loan(db, fam.id, lender="LoanZ", principal=1000, monthly_payment=100)
    # Redirect output dir so we don't touch the repo
    monkeypatch.setattr(app_config.settings, "OUTPUT_DIR", tmp_path)
    out_path = excel_export.export_monthly(db, fam.id)
    wb = load_workbook(out_path)
    assert "Loans" in wb.sheetnames
    ws = wb["Loans"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert "Lender" in header and "Monthly Payment" in header
    body = [r for r in rows[1:] if r[0] is not None]  # skip totals padding
    assert any(r[2] == "LoanZ" for r in body)


# ---------------------------------------------------------------------------
# Reminder service: compute_next_due + dedup
# ---------------------------------------------------------------------------


def test_compute_next_due_same_month():
    from datetime import date as _d
    from app.services.reminder_service import compute_next_due
    # today=May 9, due_day=10 -> next due is May 10
    assert compute_next_due(10, _d(2026, 5, 9)) == _d(2026, 5, 10)
    # today=May 10 (the due day) -> still today
    assert compute_next_due(10, _d(2026, 5, 10)) == _d(2026, 5, 10)


def test_compute_next_due_rolls_to_next_month():
    from datetime import date as _d
    from app.services.reminder_service import compute_next_due
    # today=May 11, due_day=10 -> rolled to June 10
    assert compute_next_due(10, _d(2026, 5, 11)) == _d(2026, 6, 10)
    # today=Dec 31 due_day=5 -> Jan 5 next year
    assert compute_next_due(5, _d(2026, 12, 31)) == _d(2027, 1, 5)


def test_compute_next_due_clamps_short_months():
    from datetime import date as _d
    from app.services.reminder_service import compute_next_due
    # due_day=31 in February -> Feb 28 (non-leap 2026)
    assert compute_next_due(31, _d(2026, 2, 1)) == _d(2026, 2, 28)
    # due_day=31 in April -> April 30
    assert compute_next_due(31, _d(2026, 4, 1)) == _d(2026, 4, 30)
    # day after April 30 with due_day=31 -> May 31
    assert compute_next_due(31, _d(2026, 5, 1)) == _d(2026, 5, 31)


def test_recurring_expense_crud_and_scoping():
    from app.services import recurring_expense_service as svc
    db = _make_db()
    fa = _make_family(db, "A")
    fb = _make_family(db, "B")
    a_item = svc.create_recurring(db, fa.id, name="TNB", amount=120, payment_due_day=10)
    svc.create_recurring(db, fb.id, name="Netflix", amount=55, payment_due_day=5)
    assert len(svc.list_recurring(db, fa.id)) == 1
    assert svc.get_recurring(db, fb.id, a_item.id) is None  # cross-family
    upd = svc.update_recurring(db, fa.id, a_item.id, amount=130, status="paused")
    assert upd.amount == 130 and upd.status == "paused"
    # invalid status ignored
    upd2 = svc.update_recurring(db, fa.id, a_item.id, status="garbage")
    assert upd2.status == "paused"
    # delete
    assert svc.delete_recurring(db, fa.id, a_item.id) is True
    assert svc.get_recurring(db, fa.id, a_item.id) is None


def test_recurring_due_day_clamped():
    from app.services import recurring_expense_service as svc
    db = _make_db()
    fam = _make_family(db)
    # 99 should clamp to 31 (model takes a month-agnostic 1..31)
    item = svc.create_recurring(db, fam.id, name="X", amount=10, payment_due_day=99)
    assert item.payment_due_day == 31
    item2 = svc.create_recurring(db, fam.id, name="Y", amount=10, payment_due_day=0)
    assert item2.payment_due_day == 1


def test_reminder_run_dedup(monkeypatch):
    """Calling run_for_family twice on the same day must not double-send.
    Stub send_text so we can count calls without hitting the network."""
    from datetime import date as _d
    from app.services import recurring_expense_service as svc
    from app.services import reminder_service
    from app.models import WhatsappEnrollment

    db = _make_db()
    fam = _make_family(db)
    # one enrollment so messages get a target
    db.add(WhatsappEnrollment(family_id=fam.id, whatsapp_number="60123"))
    db.commit()
    # bill due tomorrow
    today = _d(2026, 5, 9)
    svc.create_recurring(db, fam.id, name="TNB", amount=88, payment_due_day=10)

    sends = []
    monkeypatch.setattr(
        reminder_service.whatsapp_service,
        "send_text",
        lambda num, body: (sends.append((num, body)) or True),
    )

    sent1, dup1 = reminder_service.run_for_family(db, fam.id, today=today)
    assert (sent1, dup1) == (1, 0)
    assert len(sends) == 1
    assert "TOMORROW" in sends[0][1]

    # Second call same day -> all skipped as dup
    sent2, dup2 = reminder_service.run_for_family(db, fam.id, today=today)
    assert (sent2, dup2) == (0, 1)
    assert len(sends) == 1  # no new sends


def test_reminder_day_of_message(monkeypatch):
    from datetime import date as _d
    from app.services import loan_service, reminder_service
    from app.models import WhatsappEnrollment

    db = _make_db()
    fam = _make_family(db)
    db.add(WhatsappEnrollment(family_id=fam.id, whatsapp_number="60999"))
    db.commit()
    loan_service.create_loan(
        db, fam.id, lender="Maybank", principal=300000, monthly_payment=1500,
        payment_due_day=9,  # today
    )
    sends = []
    monkeypatch.setattr(
        reminder_service.whatsapp_service,
        "send_text",
        lambda num, body: (sends.append(body) or True),
    )
    sent, dup = reminder_service.run_for_family(db, fam.id, today=_d(2026, 5, 9))
    assert sent == 1 and dup == 0
    assert "TODAY" in sends[0]
    assert "Maybank" in sends[0]
    assert "RM 1500.00" in sends[0]  # MYR -> RM symbol


def test_upcoming_for_family_combines_loans_and_recurring():
    from datetime import date as _d
    from app.services import loan_service, recurring_expense_service as svc, reminder_service
    db = _make_db()
    fam = _make_family(db)
    loan_service.create_loan(db, fam.id, lender="A", principal=100, monthly_payment=10, payment_due_day=10)
    svc.create_recurring(db, fam.id, name="TNB", amount=88, payment_due_day=15)
    items = reminder_service.upcoming_for_family(db, fam.id, today=_d(2026, 5, 1), days_ahead=30)
    types = sorted([(i.target_type, i.name) for i in items])
    assert ("loan", "A") in types
    assert ("recurring", "TNB") in types
    # Sorted earliest-first
    assert items[0].due_date <= items[1].due_date


# ---------------------------------------------------------------------------
# Currency: format_money + parse_currency_hint + ask_currency step
# ---------------------------------------------------------------------------


def test_format_money_each_currency():
    from app.utils.currency import format_money
    assert format_money(88.5, "MYR") == "RM 88.50"
    assert format_money(50, "USD") == "$ 50.00"
    assert format_money(50, "SGD") == "S$ 50.00"
    assert format_money(50, "AUD") == "A$ 50.00"
    assert format_money(100, "EUR") == "€ 100.00"
    assert format_money(100, "GBP") == "£ 100.00"
    assert format_money(1000, "JPY") == "¥ 1000.00"
    assert format_money(1000, "HKD") == "HK$ 1000.00"
    # Unknown code falls back to the code itself
    assert format_money(50, "XYZ") == "RM 50.00"  # normalize → MYR


def test_parse_currency_hint_iso_codes():
    from app.utils.currency import parse_currency_hint
    assert parse_currency_hint("USD 50 lunch") == "USD"
    assert parse_currency_hint("Tesco MYR 88.50") == "MYR"
    assert parse_currency_hint("usd 50 lowercase") == "USD"  # case-insensitive code
    assert parse_currency_hint("salary 3800 SGD") == "SGD"
    assert parse_currency_hint("no currency here") is None


def test_parse_currency_hint_symbols():
    from app.utils.currency import parse_currency_hint
    # Order matters — S$ must beat $
    assert parse_currency_hint("S$50 hawker") == "SGD"
    assert parse_currency_hint("HK$200") == "HKD"
    assert parse_currency_hint("$50 lunch") == "USD"
    assert parse_currency_hint("RM88 groceries") == "MYR"
    assert parse_currency_hint("£100 london") == "GBP"
    assert parse_currency_hint("€50 vacation") == "EUR"


def test_currency_normalize():
    from app.utils.currency import normalize, is_supported
    assert normalize("usd") == "USD"
    assert normalize("xyz") == "MYR"
    assert normalize("") == "MYR"
    assert is_supported("USD") is True
    assert is_supported("XYZ") is False


def test_rule_parser_extracts_currency_hint():
    from app.services import rule_parser
    r = rule_parser.parse("USD 50 lunch")
    assert r["currency"] == "USD"
    r = rule_parser.parse("S$ 25 hawker")
    assert r["currency"] == "SGD"
    r = rule_parser.parse("plain text no money")
    assert r["currency"] == ""  # parser leaves empty so caller can default


def test_question_engine_resolves_currency():
    assert question_engine.resolve_answer("ask_currency", "A") == "MYR"
    assert question_engine.resolve_answer("ask_currency", "B") == "SGD"
    assert question_engine.resolve_answer("ask_currency", "C") == "USD"
    assert question_engine.resolve_answer("ask_currency", "Z") is None


def test_determine_next_asks_currency_after_amount():
    rec = _StubRecord(record_type="expense", amount=50, currency="")
    step, _q, _opts = question_engine.determine_next_question(rec)
    assert step == "ask_currency"


def test_determine_next_skips_currency_when_present():
    rec = _StubRecord(record_type="expense", amount=50, currency="USD")
    step, _q, _opts = question_engine.determine_next_question(rec)
    assert step == "ask_category"  # past ask_currency


def test_loan_service_currency_normalization():
    from app.services import loan_service
    db = _make_db()
    fam = _make_family(db)
    a = loan_service.create_loan(db, fam.id, lender="A", principal=100, monthly_payment=10, currency="usd")
    assert a.currency == "USD"  # normalized
    b = loan_service.create_loan(db, fam.id, lender="B", principal=100, monthly_payment=10, currency="garbage")
    assert b.currency == "MYR"  # falls back to default
    upd = loan_service.update_loan(db, fam.id, a.id, currency="sgd")
    assert upd.currency == "SGD"


def test_recurring_service_currency_normalization():
    from app.services import recurring_expense_service as svc
    db = _make_db()
    fam = _make_family(db)
    item = svc.create_recurring(db, fam.id, name="X", amount=10, payment_due_day=10, currency="eur")
    assert item.currency == "EUR"
    upd = svc.update_recurring(db, fam.id, item.id, currency="garbage")
    # garbage falls back to MYR via normalize
    assert upd.currency == "MYR"


def test_reminder_message_uses_item_currency(monkeypatch):
    from datetime import date as _d
    from app.services import loan_service, reminder_service
    from app.models import WhatsappEnrollment
    db = _make_db()
    fam = _make_family(db)
    db.add(WhatsappEnrollment(family_id=fam.id, whatsapp_number="60123"))
    db.commit()
    loan_service.create_loan(
        db, fam.id, lender="HSBC", principal=300000, monthly_payment=1500,
        currency="USD", payment_due_day=9,
    )
    sends = []
    monkeypatch.setattr(
        reminder_service.whatsapp_service,
        "send_text",
        lambda num, body: (sends.append(body) or True),
    )
    sent, _ = reminder_service.run_for_family(db, fam.id, today=_d(2026, 5, 9))
    assert sent == 1
    # Should display USD with $ symbol, not MYR/RM
    assert "$ 1500.00" in sends[0]
